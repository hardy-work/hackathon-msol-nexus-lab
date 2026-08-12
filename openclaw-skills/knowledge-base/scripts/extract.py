#!/usr/bin/env python3
"""STAGE 2 — EXTRACT (trích xuất KHAI BÁO).

originals/*.xlsx  --[ extract/<doc_id>.yml ]-->  raw/*.md + raw/*.facts.json

Nguyên tắc: script này KHÔNG ĐOÁN gì cả. Nó chỉ đọc đúng những sheet/ô/cột
được khai trong mapping. Sheet không khai = không đọc. Đoán sai cấu trúc Excel
là lỗi thầm lặng, nên ở đây không có chỗ cho việc đoán.

GATE 2 — numeric_guard(policy=ingest):
  mọi giá trị vào facts.json phải là số thật (int/float), có unit, có src trỏ về
  đúng ô trong file gốc. Vi phạm -> HALT, không ghi gì cả.
"""
import json
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl
import yaml

import numeric_guard
from numeric_guard import Halt
import document_registry
from artifact_paths import artifact_path
from openpyxl.utils import column_index_from_string as col_idx
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "raw"


# ---------------------------------------------------------------- tiện ích
def slugify(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace("Đ", "D").replace("đ", "d").lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s


# Slug của `assignee` là một giá trị DIMENSION -> theo §1.1 phải CHỌN từ danh sách đóng,
# KHÔNG được tự chế. Trước đây Stage 2 slugify() nhãn gốc -> bịa slug: `[FE]H.Anh` ra
# `fe-h-anh` (người muốn `fe-hanh`), `PhongĐT` ra `phongdt` (muốn `phong-dt`)… vì cắt biên
# camelCase/dấu chấm là quyết định người-curate, máy không suy được. Đây là gốc drift #3.
# Nay tra bảng nhãn→slug trong schema.yml (bản máy đọc của CLAUDE.md §2, do lint-contract
# canh khớp). Nhãn lạ = coverage thủng -> HALT, không lặng lẽ đẻ slug mới.
_ASSIGNEE_SLUG = None


def assignee_slug(label: str) -> str:
    global _ASSIGNEE_SLUG
    if _ASSIGNEE_SLUG is None:
        sc = yaml.safe_load((ROOT / "schema.yml").read_text(encoding="utf-8"))
        _ASSIGNEE_SLUG = {v.strip(): k for k, v in sc["dimensions"]["assignee"]["values"].items()}
    key = re.sub(r"\s+", " ", str(label)).strip()
    if key not in _ASSIGNEE_SLUG:
        raise Halt(
            f"assignee '{label}' không có trong DIMENSION assignee (schema.yml / CLAUDE.md §2). "
            f"Stage 2 KHÔNG tự chế slug (§1.1) — người thêm nhãn vào §2 + schema.yml rồi chạy lại."
        )
    return _ASSIGNEE_SLUG[key]


def cell_str(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# GATE 2 dùng chung một cơ chế với GATE 4 — xem scripts/numeric_guard.py
def numeric_guard_ingest(name, value, unit, src):
    return numeric_guard.check(policy="ingest", name=name, value=value, unit=unit, src=src)


# ---------------------------------------------------------- ba loại nguồn
def do_vocabulary(ws, spec, sheet):
    """Sheet Config -> danh sách giá trị DIMENSION. Đây là nguồn của từ vựng đóng.
    Mỗi giá trị kèm toạ độ ô nguồn (`vocabulary_cells`) — provenance cấp ô, để
    `audit_raw.py` đếm được (nó nhặt mọi dict có khoá `src`) và để mỗi giá trị DIMENSION
    truy được về đúng ô như MEASURE."""
    vocab, cells, lines = {}, [], []
    for field, cfg in spec["columns"].items():
        a, b = cfg["range"].split(":")
        c0, r0 = re.match(r"([A-Z]+)(\d+)", a).groups()
        _, r1 = re.match(r"([A-Z]+)(\d+)", b).groups()
        seen = []
        for r in range(int(r0), int(r1) + 1):
            v = cell_str(ws.cell(r, col_idx(c0)).value)
            if v and v not in seen:
                seen.append(v)
                cells.append({"field": field, "value": v, "src": f"{sheet}!{c0}{r}"})
        vocab[field] = seen
        lines.append(f"## {field}  ({cfg['range']}, {len(seen)} giá trị)\n")
        lines += [f"- `{v}`" for v in seen]
        lines.append("")
    return "\n".join(lines), {"vocabulary": vocab, "vocabulary_cells": cells}, len(vocab)


def do_cells(ws, spec, sheet):
    """Ô rời rạc -> facts.json. Mỗi ô khai đích danh."""
    facts, lines = {}, ["| khoá | nhãn | ô | giá trị | đơn vị |", "|---|---|---|---|---|"]
    for key, cfg in spec["facts"].items():
        addr = cfg["cell"]
        val = ws[addr].value
        src = f"{sheet}!{addr}"
        facts[key] = numeric_guard_ingest(key, val, cfg["unit"], src)
        lines.append(f"| `{key}` | {cfg.get('label','')} | `{addr}` | {val} | {cfg['unit']} |")
    # ghi `sheet` để Gate 4 theo-ngữ-cảnh khớp được số này khi câu trả lời trích ô của nó
    # (cite kiểu '… → 1.Summary Project!J17'). KHÔNG vào doc_cell (không có `rows`).
    return "\n".join(lines), {"facts": facts, "sheet": sheet}, len(facts)


def _fmt_cell(cell):
    """Giá trị ô -> chuỗi để đọc. Ngày -> YYYY-MM-DD; ô phần trăm -> 'x%';
    số nguyên -> không đuôi .0. Trả None nếu ô rỗng."""
    v = cell.value
    if v is None or v == "":
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, float):
        if "%" in (cell.number_format or ""):
            return f"{v * 100:g}%"
        if v.is_integer():
            return str(int(v))
        return str(round(v, 6))
    return str(v).strip().replace("\n", " ")


def do_fulltext(ws, spec, sheet):
    """LUỒNG VĂN — nạp TRỌN sheet: mọi ô không rỗng, KHÔNG bỏ cột nào.

    Dùng cho tài liệu/bảng mà ta không (và không nên) khai từng ô. Ngược hẳn với
    do_table (lỗ khoá, chỉ cột đã khai): đây là vòi rồng. Vì không rút MEASURE nào
    nên không đi qua numeric_guard(ingest); nhưng MỌI số vẫn được ghi kèm ô nguồn
    (`cells`) để Gate 4 truy được về raw/ — số trong fulltext vẫn không phải số bịa."""
    max_rows = spec.get("max_rows", 2000)
    max_col = ws.max_column or 1
    lines, cells, nrows, blank, r = [], {}, 0, 0, 1
    while blank < 30 and r <= max_rows:
        parts, row_has = [], False
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            s = _fmt_cell(cell)
            if s is None:
                continue
            row_has = True
            col = get_column_letter(c)
            parts.append(f"{col}={s}")
            v = cell.value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                coord = f"{col}{r}"
                cells[coord] = {"value": v, "src": f"{sheet}!{coord}"}
        if row_has:
            lines.append(f"[r{r}] " + " | ".join(parts))
            nrows += 1
            blank = 0
        else:
            blank += 1
        r += 1
    body = (f"_Nạp TRỌN sheet `{sheet}` — {nrows} dòng có dữ liệu, {len(cells)} ô số, "
            f"KHÔNG bỏ cột nào._\n\n" + "\n".join(lines))
    payload = {"kind": "fulltext", "sheet": sheet, "cells": cells}
    return body, payload, nrows


def scan_headers(ws, header_rows, max_col, min_col=1):
    """Nhãn cột. Nhiều dòng tiêu đề (gộp ô) thì nối lại: PLAN + Start Date."""
    headers = {}
    for c in range(min_col, max_col + 1):
        parts = [_fmt_cell(ws.cell(hr, c)) for hr in header_rows]
        lab = " ".join(p for p in parts if p).strip()
        headers[get_column_letter(c)] = lab or get_column_letter(c)
    return headers


def col_range(spec, ws):
    """`cols: "D:F"` -> (4, 6). Không khai thì lấy trọn sheet. Cần cho sheet chứa
    NHIỀU bảng rời nhau (Config có cả `1. Configure` lẫn `2. Holiday`)."""
    if not spec.get("cols"):
        return 1, (ws.max_column or 1)
    a, b = spec["cols"].split(":")
    return col_idx(a), col_idx(b)


def scan_rows(ws, sheet, headers, start, max_col, skip_rows=(), min_col=1):
    """Quét MỌI ô của MỌI dòng dữ liệu -> record kèm ô nguồn. Dùng chung cho
    `kind: rows` và `kind: table` — bảng nào nạp vào cũng phải tra được ở bậc 1,
    kể cả cột không khai làm MEASURE/DIMENSION (ví dụ `Note`)."""
    rows, numeric, blank, r = [], {}, 0, start
    while blank < 30 and r <= (ws.max_row or start):
        if r in skip_rows:
            r += 1
            continue
        cells = {}
        for c in range(min_col, max_col + 1):
            s = _fmt_cell(ws.cell(r, c))
            if s is None:
                continue
            col = get_column_letter(c)
            v = ws.cell(r, c).value
            src = f"{sheet}!{col}{r}"
            cells[col] = {
                "header": headers[col],
                "value": s,
                "value_num": v if isinstance(v, (int, float)) and not isinstance(v, bool) else None,
                "src": src,
            }
            for tok in re.findall(r"\d+(?:[.,]\d+)?", s):
                normalized = tok.replace(",", ".")
                numeric[f"{col}{r}:{tok}"] = {
                    "value": float(normalized) if "." in normalized else int(normalized),
                    "text": tok, "src": src}
        if cells:
            rows.append({"row": r, "cells": cells})
            blank = 0
        else:
            blank += 1
        r += 1
    return rows, numeric


def do_rows(ws, spec, sheet):
    """BẢNG BÁN CẤU TRÚC -> mỗi DÒNG một record, GIỮ ô nguồn. Đây là đường ĐÚNG FLOW
    cho dữ liệu bảng (.xlsx): số vào duckdb bậc 1, truy được về ô — KHÔNG coi như văn
    xuôi, KHÔNG để LLM đọc raw. Khác do_table (gộp theo người) và do_fulltext (prose).

    Khai `header_row` (lấy nhãn cột) + `data_start_row`. Không đoán cấu trúc."""
    hr = spec["header_row"]
    start = spec["data_start_row"]
    min_col, max_col = col_range(spec, ws)
    headers = scan_headers(ws, [hr], max_col, min_col)

    # numeric_cells: MỌI số xuất hiện trong MỌI ô (kể cả số lẫn trong văn xuôi như
    # "24 items"), ghi kèm ô nguồn. Đây là để Gate 4(answer) whitelist được — số trong
    # câu trả lời trích verbatim từ ô raw thì truy được nguồn, không phải số bịa.
    rows, numeric = scan_rows(ws, sheet, headers, start, max_col, min_col=min_col)

    lines = [f"_Bảng `{sheet}` — {len(rows)} dòng, header dòng {hr}, dữ liệu từ dòng {start}._", ""]
    for rec in rows:
        parts = [f"{cell['header']}={cell['value']}" for cell in rec["cells"].values()]
        lines.append(f"[r{rec['row']}] " + " | ".join(parts))
    payload = {"kind": "rows", "sheet": sheet, "headers": headers,
               "rows": rows, "numeric_cells": numeric}
    return "\n".join(lines), payload, len(rows)


def do_table(ws, spec, sheet, raw_id):
    """Bảng task -> các dòng + tổng hợp theo assignee/role."""
    cols = spec["columns"]
    dim_cols = {f: c["dimension"] for f, c in cols.items() if "dimension" in c}
    num_cols = {f: c["unit"] for f, c in cols.items() if "unit" in c}
    start = spec["data_start_row"]

    rows = []
    r = start
    blank = 0
    while blank < 20 and r < ws.max_row + 2:
        rec, has = {}, False
        for f, cfg in cols.items():
            v = ws.cell(r, col_idx(cfg["col"])).value
            rec[f] = v
            if v not in (None, ""):
                has = True
        if has:
            rec["_row"] = r
            rows.append(rec)
            blank = 0
        else:
            blank += 1
        r += 1

    # tổng hợp theo assignee — đây là thứ trang entities/ sẽ trỏ tới
    by_assignee = {}
    for rec in rows:
        who = cell_str(rec.get("assignee"))
        if not who:
            continue
        k = assignee_slug(who)
        agg = by_assignee.setdefault(
            k, {"label": who, "task_count": 0, "rows": [], "roles": []}
        )
        agg["task_count"] += 1
        agg["rows"].append(rec["_row"])
        role = cell_str(rec.get("role"))
        if role and role not in agg["roles"]:
            agg["roles"].append(role)
        for f, unit in num_cols.items():
            v = rec.get(f)
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                agg[f] = round(agg.get(f, 0.0) + float(v), 4)

    facts = {}
    for k, agg in by_assignee.items():
        rr = agg["rows"]
        span = f"{sheet}!{cols['assignee']['col']}{rr[0]}:{cols['assignee']['col']}{rr[-1]}"
        facts[k] = {
            "label": agg["label"],
            "roles": agg["roles"],
            "task_count": numeric_guard_ingest(
                f"{k}.task_count", agg["task_count"], "task", span
            ),
        }
        for f, unit in num_cols.items():
            if f in agg:
                facts[k][f] = numeric_guard_ingest(
                    f"{k}.{f}", agg[f], unit,
                    f"{sheet}!{cols[f]['col']}{rr[0]}:{cols[f]['col']}{rr[-1]}",
                )

    # giá trị DIMENSION thực tế gặp trong sheet — Gate 3a sẽ đối chiếu với CLAUDE.md
    observed = {
        dim: sorted({cell_str(rec.get(f)) for rec in rows if cell_str(rec.get(f))})
        for f, dim in dim_cols.items()
    }

    _ = raw_id
    hdr = [f for f in cols if f != "_row"]
    lines = ["| dòng | " + " | ".join(hdr) + " |", "|" + "---|" * (len(hdr) + 1)]
    for rec in rows:
        lines.append(
            f"| {rec['_row']} | " + " | ".join(cell_str(rec.get(f)) for f in hdr) + " |"
        )
    lines += ["", f"_{len(rows)} dòng, dữ liệu từ row {start} (đã bỏ dòng tổng nhỏ row {spec.get('subtotal_row')})._"]

    # Ngoài phần gộp MEASURE ở trên, sheet task còn đi tiếp vào BẬC 1 dạng row-record
    # phủ MỌI cột (kể cả cột không khai MEASURE/DIMENSION như `Note`). Nạp cái gì thì
    # phải tra được cái đó — mapping lỗ khoá chỉ quyết định cột nào thành MEASURE,
    # không được quyết định cột nào biến mất khỏi kho.
    max_col = ws.max_column or 1
    headers = scan_headers(ws, spec["header_rows"], max_col)
    skip = {spec["subtotal_row"]} if spec.get("subtotal_row") else set()
    cell_rows, numeric = scan_rows(ws, sheet, headers, start, max_col, skip_rows=skip)
    payload = {"facts": facts, "observed_dimensions": observed, "sheet": sheet,
               "headers": headers, "rows": cell_rows, "numeric_cells": numeric}
    return "\n".join(lines), payload, len(rows)


def do_rollup(spec, vocab_facts, document):
    """Cộng nhiều nguồn table đã sinh -> một nguồn tổng hợp.
    Người khai trong Config nhưng 0 task VẪN có mục ở đây với task_count=0 —
    đây chính là bằng chứng để bậc 1 nói 'CHẮC CHẮN KHÔNG'."""
    agg, srcs = {}, {}
    for rid in spec["over"]:
        f = artifact_path(ROOT, document, rid, "facts")
        if not f.exists():
            raise Halt(f"rollup cần {f.name} nhưng chưa có — khai nó TRƯỚC rollup trong mapping")
        for k, v in json.loads(f.read_text(encoding="utf-8"))["facts"].items():
            a = agg.setdefault(k, {"label": v["label"], "roles": []})
            for r in v.get("roles", []):
                if r not in a["roles"]:
                    a["roles"].append(r)
            for field in spec["sum"]:
                if field in v:
                    a.setdefault(field, {"value": 0.0, "unit": v[field]["unit"]})
                    tot = a[field]["value"] + v[field]["value"]
                    a[field]["value"] = int(tot) if field == "task_count" else round(tot, 4)
                    srcs.setdefault((k, field), []).append(v[field]["src"])

    # người khai trong Config nhưng không xuất hiện ở sprint nào -> 0, có chủ đích
    declared = vocab_facts["vocabulary"]["assignee"]
    zero = []
    for label in declared:
        k = assignee_slug(label)
        if k not in agg:
            zero.append(label)
            agg[k] = {"label": label, "roles": [],
                      **{f: {"value": 0, "unit": "task" if f == "task_count" else "hour"}
                         for f in spec["sum"]}}

    facts = {}
    for k, a in agg.items():
        node = {"label": a["label"], "roles": a["roles"],
                "in_config": True, "appears_in_sprints": a["label"] not in zero}
        for field in spec["sum"]:
            src = srcs.get((k, field))
            node[field] = numeric_guard_ingest(
                f"{k}.{field}", a[field]["value"], a[field]["unit"],
                " + ".join(src) if src else "Config!B3:B11 (khai trong Config, 0 dòng ở Sprint 0-7)",
            )
        facts[k] = node

    # người có task nhưng KHÔNG khai trong Config -> coverage thủng, phải HALT
    undeclared = [a["label"] for k, a in agg.items() if a["label"] not in declared]
    if undeclared:
        raise Halt(f"assignee có task nhưng không khai trong Config: {undeclared} -> coverage thủng")

    lines = ["| người | vai trò | task | estimate (h) | actual (h) |", "|---|---|---|---|---|"]
    for k, n in sorted(facts.items(), key=lambda x: -x[1]["task_count"]["value"]):
        lines.append(f"| {n['label']} | {', '.join(n['roles']) or '—'} | "
                     f"{n['task_count']['value']} | {n['estimate_h']['value']} | {n['actual_h']['value']} |")
    lines += ["", f"_{len(facts)} người khai trong Config · {len(zero)} người 0 task: "
                  f"{', '.join(zero) or 'không có'}_",
              "", "_Không có người nào làm task mà không được khai trong Config → phạm vi kín._"]
    return "\n".join(lines), {"facts": facts, "zero_task": zero}, len(facts)


# ------------------------------------------------------------------ chạy
def write_raw(raw_id, doc_id, sheet, kind, body, payload, document, suffix=".md"):
    artifact_kind = "fulltext" if suffix == ".fulltext.md" else "md"
    md = artifact_path(ROOT, document, raw_id, artifact_kind)
    fj = artifact_path(ROOT, document, raw_id, "facts")
    md.parent.mkdir(parents=True, exist_ok=True)
    payload = {**payload, "doc_id": doc_id, "version": int(document["version"])}
    md.write_text(
        f"---\n"
        f"raw_id: {raw_id}\ndoc_id: {doc_id}\nsheet: \"{sheet}\"\nkind: {kind}\n"
        f"version: {int(document['version'])}\n"
        f"generated_by: scripts/extract.py\n"
        f"# TẦNG 2 — MÁY SINH. KHÔNG SỬA TAY. Muốn đổi thì sửa extract/*.yml rồi chạy lại.\n"
        f"---\n\n# {raw_id}\n\nNguồn: `{sheet}` trong `originals/`\n\n{body}\n",
        encoding="utf-8",
    )
    fj.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return md, fj


def main(mapping_path):
    spec = yaml.safe_load(Path(mapping_path).read_text(encoding="utf-8"))
    doc_id = spec["doc_id"]
    document = document_registry.current(doc_id, ROOT)
    wb = openpyxl.load_workbook(ROOT / spec["original"], data_only=True)
    RAW.mkdir(exist_ok=True)

    declared = set()
    made = []
    vocab_facts = None

    for src in spec["sources"]:
        if src["kind"] == "rollup":
            if vocab_facts is None:
                raise Halt("rollup cần nguồn kind=vocabulary được khai trước nó")
            body, payload, n = do_rollup(src, vocab_facts, document)
            write_raw(src["raw_id"], doc_id, "(tổng hợp 8 sprint)", "rollup", body, payload, document)
            made.append((src["raw_id"], "-", "rollup", n))
            print(f"  ✓ {src['raw_id']:20s} {'rollup':11s} {n:4d}  <- {len(src['over'])} nguồn")
            continue

        sheets = (
            [(src["raw_id"], src["sheet"])]
            if "raw_id" in src
            else [(src["raw_id_pattern"].format(n=s["n"]), s["sheet"]) for s in src["sheets"]]
        )
        for raw_id, sheet in sheets:
            declared.add(sheet)
            if sheet not in wb.sheetnames:
                raise Halt(f"mapping khai sheet '{sheet}' nhưng file gốc không có")
            ws = wb[sheet]
            kind = src["kind"]
            fn = {"vocabulary": do_vocabulary, "cells": do_cells}.get(kind)
            if fn:
                body, payload, n = fn(ws, src, sheet)
            elif kind == "table":
                body, payload, n = do_table(ws, src, sheet, raw_id)
            elif kind == "fulltext":
                body, payload, n = do_fulltext(ws, src, sheet)
            elif kind == "rows":
                body, payload, n = do_rows(ws, src, sheet)
            else:
                raise Halt(f"kind '{kind}' không hỗ trợ")
            if kind == "vocabulary":
                vocab_facts = payload
            suffix = ".fulltext.md" if kind == "fulltext" else ".md"
            md, fj = write_raw(raw_id, doc_id, sheet, kind, body, payload, document, suffix)
            made.append((raw_id, sheet, kind, n))
            print(f"  ✓ {raw_id:20s} {kind:11s} {n:4d}  <- {sheet}")

    # mọi sheet phải hoặc được khai, hoặc bị loại CÓ LÝ DO. Không có vùng xám.
    excluded = {e["sheet"] for e in spec.get("excluded_sheets", [])}
    unaccounted = set(wb.sheetnames) - declared - excluded
    if unaccounted:
        raise Halt(
            "sheet không khai và cũng không loại: "
            + ", ".join(sorted(unaccounted))
            + "\n  -> thêm vào sources hoặc excluded_sheets (kèm reason) rồi chạy lại."
        )

    print(f"\n{len(made)} nguồn raw/ · {len(declared)} sheet trích · {len(excluded)} sheet loại có lý do")
    return made


if __name__ == "__main__":
    m = sys.argv[1] if len(sys.argv) > 1 else "extract/handy-schedule-v2.1.yml"
    print(f"STAGE 2 · EXTRACT · {m}\n")
    try:
        main(ROOT / m)
    except Halt as e:
        print(f"\n✗ HALT — {e}", file=sys.stderr)
        sys.exit(1)
