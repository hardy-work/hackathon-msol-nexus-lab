#!/usr/bin/env python3
"""KIỂM KÊ originals/ -> raw/ : còn ô nào của file gốc CHƯA vào raw/ không?

Không tin lời, đo bằng ô. Mở lại file gốc, liệt kê MỌI ô không rỗng, rồi đối chiếu
với tập toạ độ ô mà raw/*.facts.json đã đăng ký (`src` = "Sheet!A12"). Chênh lệch
chính là phần dữ liệu đang nằm ngoài kho — dù file đã "nạp rồi".

  python3 scripts/audit_raw.py          # bảng tổng hợp
  python3 scripts/audit_raw.py -v       # kèm ví dụ ô bị bỏ
"""
import json
import sys
from pathlib import Path

import openpyxl
import yaml
from openpyxl.utils import get_column_letter

import document_registry
from artifact_paths import payload_is_current

ROOT = Path(__file__).resolve().parent.parent
G, R, Y, D, OFF = "\033[32m", "\033[31m", "\033[33m", "\033[2m", "\033[0m"


def covered_cells():
    """{sheet: {toạ_độ}} — mọi ô raw/ đã đăng ký kèm src."""
    cov = {}

    def walk(node):
        if isinstance(node, dict):
            src = node.get("src")
            if isinstance(src, str) and "!" in src:
                sheet, ref = src.rsplit("!", 1)
                cov.setdefault(sheet, set()).add(ref)
            for v in node.values():
                walk(v)
        elif isinstance(node, list):
            for v in node:
                walk(v)

    versions = document_registry.current_versions(ROOT)
    for p in sorted((ROOT / "raw").glob("*.facts.json")):
        payload = json.loads(p.read_text(encoding="utf-8"))
        if payload_is_current(payload, ROOT, versions, path=p):
            walk(payload)
    return cov


def expand(ref):
    """'A3:A16' -> {'A3',...}. Ô đơn thì trả chính nó."""
    if ":" not in ref:
        return {ref}
    import re
    a, b = ref.split(":")
    m1, m2 = re.match(r"([A-Z]+)(\d+)", a), re.match(r"([A-Z]+)(\d+)", b)
    if not (m1 and m2):
        return {ref}
    from openpyxl.utils import column_index_from_string as ci
    out = set()
    for c in range(ci(m1.group(1)), ci(m2.group(1)) + 1):
        for r in range(int(m1.group(2)), int(m2.group(2)) + 1):
            out.add(f"{get_column_letter(c)}{r}")
    return out


def main():
    verbose = "-v" in sys.argv
    cov_raw = covered_cells()
    cov = {s: set().union(*(expand(r) for r in refs)) if refs else set()
           for s, refs in cov_raw.items()}

    mapping = {"original": "originals/nexus-plan.xlsx", "excluded_sheets": []}
    excluded = {e["sheet"]: e["reason"] for e in mapping.get("excluded_sheets", [])}

    wb = openpyxl.load_workbook(ROOT / mapping["original"], data_only=True)
    print(f"{D}nguồn: {mapping['original']}{OFF}\n")
    print(f"{'sheet':28s} {'ô có data':>10s} {'đã vào raw':>11s} {'bỏ sót':>8s}  ghi chú")

    tot_all = tot_cov = 0
    for name in wb.sheetnames:
        ws = wb[name]
        cells = {c.coordinate for row in ws.iter_rows() for c in row
                 if c.value not in (None, "")}
        have = cells & cov.get(name, set())
        miss = cells - cov.get(name, set())
        tot_all += len(cells)
        tot_cov += len(have)

        if name in excluded:
            note = f"{Y}CỐ Ý LOẠI{OFF} — {excluded[name][:44]}"
        elif not miss:
            note = f"{G}đủ{OFF}"
        else:
            rows_missed = sorted({int("".join(ch for ch in m if ch.isdigit())) for m in miss})
            note = f"{R}thiếu{OFF} dòng {rows_missed[0]}–{rows_missed[-1]}"
        pct = f"{len(have) / len(cells):.0%}" if cells else "—"
        print(f"{name:28s} {len(cells):10d} {pct:>11s} {len(miss):8d}  {note}")
        if verbose and miss and name not in excluded:
            for m in sorted(miss)[:5]:
                v = str(ws[m].value)[:60].replace("\n", " ")
                print(f"{D}      {m:>6s} = {v}{OFF}")

    print(f"\n{'TỔNG':28s} {tot_all:10d} {tot_cov / tot_all:>11.0%} {tot_all - tot_cov:8d}")
    print(f"\n{D}Ô 'bỏ sót' KHÔNG tự động là lỗi: dòng tiêu đề, ô trang trí, sheet cố ý loại\n"
          f"đều nằm ở đây. Nhưng mọi ô có NỘI DUNG THẬT mà bị bỏ đều là dữ liệu\n"
          f"người dùng hỏi sẽ không bao giờ ra. Xem bằng -v.{OFF}")


if __name__ == "__main__":
    main()
