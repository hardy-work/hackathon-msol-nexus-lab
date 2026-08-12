#!/usr/bin/env python3
"""STAGE 2 — deterministic full-workbook extraction for generic XLSX uploads.

This lane does not guess a schema or promote arbitrary columns to dimensions.
It preserves every non-empty cell in raw Markdown, keeps numeric cell locators
in a facts JSON payload, and leaves semantic interpretation to the reviewed
wiki stage.  The original workbook remains the immutable source.
"""
from __future__ import annotations

import json
import sys
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
import artifact_paths  # noqa: E402
import document_registry  # noqa: E402
from spreadsheet_reader import iter_row_pairs  # noqa: E402


ROOT = Path(__file__).resolve().parent.parent


def _value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _display(value: Any) -> str:
    converted = _value(value)
    if converted is None or converted == "":
        return ""
    if isinstance(converted, float) and converted.is_integer():
        return str(int(converted))
    return str(converted).replace("\r\n", "\n").replace("\r", "\n").replace("|", "\\|")


def extract_one(root: Path, document: dict[str, Any]) -> tuple[Path, Path, int]:
    doc_id = str(document["doc_id"])
    original = root / str(document["original"])
    if not original.is_file():
        raise FileNotFoundError(f"thiếu original: {document['original']}")
    if str(document.get("kind")) != "xlsx":
        raise ValueError(f"spreadsheet extractor chỉ nhận xlsx, nhận {document.get('kind')}")

    formulas = openpyxl.load_workbook(original, data_only=False, read_only=True)
    values = openpyxl.load_workbook(original, data_only=True, read_only=True)
    sheets: list[dict[str, Any]] = []
    numeric_cells: dict[str, dict[str, Any]] = {}
    markdown: list[str] = [
        f"# {doc_id}",
        "",
        f"Nguồn: `{document['original']}` · version `{document['version']}`.",
        "Toàn bộ ô không rỗng được giữ theo sheet và toạ độ. Không tự suy ra schema.",
        "",
    ]
    total_cells = 0
    try:
        for formula_sheet in formulas.worksheets:
            value_sheet = values[formula_sheet.title]
            rows: list[dict[str, Any]] = []
            sheet_lines = [f"## Sheet `{formula_sheet.title}`", ""]
            max_row = max(formula_sheet.max_row or 0, value_sheet.max_row or 0)
            max_col = max(formula_sheet.max_column or 0, value_sheet.max_column or 0)
            for row_no, formula_row, value_row in iter_row_pairs(
                formula_sheet, value_sheet
            ):
                width = max(len(formula_row), len(value_row))
                max_row = max(max_row, row_no)
                max_col = max(max_col, width)
                cells: dict[str, dict[str, Any]] = {}
                parts: list[str] = []
                for offset in range(width):
                    formula_cell = (
                        formula_row[offset] if offset < len(formula_row) else None
                    )
                    value_cell = value_row[offset] if offset < len(value_row) else None
                    if formula_cell is None and value_cell is None:
                        continue
                    raw = formula_cell.value if formula_cell is not None else None
                    evaluated = value_cell.value if value_cell is not None else None
                    formula = raw if isinstance(raw, str) and raw.startswith("=") else None
                    value = _value(evaluated if formula and evaluated is not None else raw)
                    if value is None or value == "":
                        continue
                    col = get_column_letter(offset + 1)
                    src = f"{formula_sheet.title}!{col}{row_no}"
                    cell_data: dict[str, Any] = {
                        "header": col,
                        "value": _display(value),
                        "src": src,
                    }
                    if formula:
                        cell_data["formula"] = formula
                    if isinstance(evaluated, (int, float)) and not isinstance(
                        evaluated, bool
                    ):
                        cell_data["value_num"] = evaluated
                        numeric_cells[src] = {"value": evaluated, "src": src}
                    cells[col] = cell_data
                    formula_note = f" [formula: {formula}]" if formula else ""
                    parts.append(f"{src}={_display(value)}{formula_note}")
                    total_cells += 1
                if cells:
                    rows.append({"row": row_no, "cells": cells})
                    sheet_lines.append(f"[r{row_no}] " + " | ".join(parts))
            markdown.extend(sheet_lines)
            markdown.append("")
            sheets.append({
                "sheet": formula_sheet.title,
                "rows": rows,
                "max_row": max_row,
                "max_column": max_col,
            })
    finally:
        formulas.close()
        values.close()

    raw_path = artifact_paths.artifact_path(root, document, doc_id, "md")
    facts_path = artifact_paths.artifact_path(root, document, doc_id, "facts")
    raw_path.parent.mkdir(parents=True, exist_ok=True)
    raw_path.write_text(
        "---\n"
        f"raw_id: {doc_id}\n"
        f"doc_id: {doc_id}\n"
        f"version: {int(document['version'])}\n"
        "kind: workbook\n"
        f"source_file: {document['original']}\n"
        "generated_by: scripts/extract_spreadsheet.py\n"
        "---\n\n" + "\n".join(markdown).rstrip() + "\n",
        encoding="utf-8",
    )
    facts_path.write_text(json.dumps({
        "doc_id": doc_id,
        "version": int(document["version"]),
        "kind": "workbook",
        "source_file": document["original"],
        "sheets": sheets,
        "numeric_cells": numeric_cells,
    }, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return raw_path, facts_path, total_cells


def main(argv: list[str] | None = None) -> int:
    args = list(argv or sys.argv[1:])
    doc_id = args[args.index("--doc") + 1] if "--doc" in args else None
    if not doc_id:
        print("Dùng: python3 scripts/extract_spreadsheet.py --doc <doc_id>", file=sys.stderr)
        return 2
    try:
        document = document_registry.current(doc_id, ROOT)
        raw, facts, cells = extract_one(ROOT, document)
    except (OSError, KeyError, ValueError) as exc:
        print(f"✗ {doc_id}: {exc}", file=sys.stderr)
        return 1
    print(f"✓ {doc_id} → {raw.relative_to(ROOT)} + {facts.relative_to(ROOT)} · {cells} ô")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
