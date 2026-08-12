#!/usr/bin/env python3
"""Build a deterministic, provider-neutral review artifact for an ingest proposal.

The JSON output is the contract a NexusBot Google Sheet/Doc publisher consumes.
The local Markdown copy is useful for CLI review and offline tests.  Neither
artifact replaces the original upload: the original file, hash and cell
locators remain the provenance source for Project Knowledge.
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import json
import mimetypes
import sys
from pathlib import Path
from typing import Any

import openpyxl

from spreadsheet_reader import iter_row_pairs


SCHEMA = "nexus-project-knowledge/review-artifact/v1"


def _json_value(value: Any) -> Any:
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _cell_value(value: Any) -> Any:
    converted = _json_value(value)
    if isinstance(converted, str):
        return converted.replace("\r\n", "\n").replace("\r", "\n")
    return converted


def _workbook(path: Path) -> list[dict[str, Any]]:
    # Read the workbook twice: formulas are retained for audit while the
    # calculated/displayed value is what the reviewer sees and what the numeric
    # extractor later uses under the existing data_only=True contract.
    formulas = openpyxl.load_workbook(path, data_only=False, read_only=True)
    values = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets: list[dict[str, Any]] = []
    try:
        for formula_sheet in formulas.worksheets:
            value_sheet = values[formula_sheet.title]
            cells: list[dict[str, Any]] = []
            max_row = max(formula_sheet.max_row or 0, value_sheet.max_row or 0)
            max_col = max(formula_sheet.max_column or 0, value_sheet.max_column or 0)
            for row, formula_row, value_row in iter_row_pairs(
                formula_sheet, value_sheet
            ):
                width = max(len(formula_row), len(value_row))
                max_row = max(max_row, row)
                max_col = max(max_col, width)
                for offset in range(width):
                    formula_cell = (
                        formula_row[offset] if offset < len(formula_row) else None
                    )
                    value_cell = value_row[offset] if offset < len(value_row) else None
                    raw = formula_cell.value if formula_cell is not None else None
                    evaluated = value_cell.value if value_cell is not None else None
                    if raw in (None, "") and evaluated in (None, ""):
                        continue
                    cell = formula_cell if raw not in (None, "") else value_cell
                    formula = raw if isinstance(raw, str) and raw.startswith("=") else None
                    displayed = evaluated if formula else raw
                    cells.append({
                        "address": cell.coordinate,
                        "value": _cell_value(displayed),
                        "formula": formula,
                        "data_type": cell.data_type,
                        "number_format": cell.number_format,
                        "source": f"{formula_sheet.title}!{cell.coordinate}",
                    })
            sheets.append({
                "name": formula_sheet.title,
                "max_row": max_row,
                "max_column": max_col,
                # Read-only worksheets do not expose merged_cells in some
                # openpyxl versions; cell values/formulas remain complete.
                "merged_ranges": [str(item) for item in
                                   getattr(getattr(formula_sheet, "merged_cells", None),
                                           "ranges", [])],
                "cells": cells,
            })
    finally:
        formulas.close()
        values.close()
    return sheets


def _csv_file(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        for row_no, row in enumerate(csv.reader(stream), start=1):
            rows.append({
                "row": row_no,
                "values": [_cell_value(value) for value in row],
                "source": f"row {row_no}",
            })
    return rows


def build(source: Path, *, proposal_id: str = "") -> dict[str, Any]:
    source = source.expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(source)
    suffix = source.suffix.casefold()
    kind = {
        ".xlsx": "xlsx", ".csv": "csv", ".md": "text/markdown",
        ".markdown": "text/markdown",
    }.get(suffix, mimetypes.guess_type(source.name)[0] or "binary")
    artifact: dict[str, Any] = {
        "schema": SCHEMA,
        "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds"),
        "proposal_id": proposal_id,
        "source": {
            "name": source.name,
            "path": str(source),
            "kind": kind,
            "size": source.stat().st_size,
            "sha256": _sha256(source),
        },
        "review": {
            "source_of_truth": "uploaded original + raw extracted artifacts",
            "google_publisher": "consumer-defined; this artifact contains no Google credential",
        },
    }
    if suffix == ".xlsx":
        artifact["workbook"] = {"sheets": _workbook(source)}
    elif suffix == ".csv":
        artifact["csv"] = {"rows": _csv_file(source)}
    elif suffix in {".md", ".markdown"}:
        artifact["markdown"] = {"text": source.read_text(encoding="utf-8")}
    else:
        artifact["unsupported_preview"] = {
            "message": "Review artifact mới giữ metadata; extractor tương ứng phải chạy trước publish.",
        }
    return artifact


def _md(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("|", "\\|").replace("\n", "<br>")


def to_markdown(artifact: dict[str, Any]) -> str:
    source = artifact["source"]
    lines = [
        "# Ingest review artifact",
        "",
        f"- Proposal: `{artifact.get('proposal_id') or '—'}`",
        f"- Source: `{source['name']}`",
        f"- Kind: `{source['kind']}`",
        f"- Size: `{source['size']}` bytes",
        f"- SHA-256: `{source['sha256']}`",
        "- Nguồn sự thật: file gốc và raw artifact; bản này chỉ dùng để review.",
        "",
    ]
    workbook = artifact.get("workbook")
    if workbook:
        for sheet in workbook["sheets"]:
            lines += [f"## Sheet: {sheet['name']}", "",
                      f"{len(sheet['cells'])} ô có dữ liệu.", "",
                      "| Ô | Giá trị hiển thị | Công thức | Nguồn |",
                      "|---|---|---|---|"]
            for cell in sheet["cells"]:
                lines.append("| " + " | ".join([
                    _md(cell["address"]), _md(cell["value"]),
                    _md(cell["formula"]), _md(cell["source"]),
                ]) + " |")
            lines.append("")
    csv_data = artifact.get("csv")
    if csv_data:
        lines += ["## CSV", ""]
        for row in csv_data["rows"]:
            lines.append(f"- `{row['source']}`: " + " | ".join(_md(v) for v in row["values"]))
        lines.append("")
    markdown = artifact.get("markdown")
    if markdown:
        lines += ["## Markdown source", "", markdown["text"], ""]
    if artifact.get("unsupported_preview"):
        lines += ["## Preview limitation", "", artifact["unsupported_preview"]["message"], ""]
    return "\n".join(lines)


def write_bundle(artifact: dict[str, Any], destination: Path) -> dict[str, Any]:
    destination = destination.expanduser().resolve()
    destination.mkdir(parents=True, exist_ok=True)
    json_path = destination / "review-artifact.json"
    markdown_path = destination / "review-artifact.md"
    json_path.write_text(json.dumps(artifact, ensure_ascii=False, indent=2) + "\n",
                         encoding="utf-8")
    markdown_path.write_text(to_markdown(artifact), encoding="utf-8")
    return {
        "schema": SCHEMA,
        "path": str(destination),
        "json_path": str(json_path),
        "markdown_path": str(markdown_path),
        "sha256": artifact["source"]["sha256"],
        "cell_count": sum(len(s["cells"]) for s in artifact.get("workbook", {}).get("sheets", [])),
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build deterministic ingest review artifact")
    parser.add_argument("--file", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--proposal-id", default="")
    args = parser.parse_args(argv)
    try:
        artifact = build(args.file, proposal_id=args.proposal_id)
        result = write_bundle(artifact, args.output)
    except (OSError, ValueError, KeyError) as exc:
        print(f"✗ {exc}", file=sys.stderr)
        return 1
    print(json.dumps({"artifact": result, "preview": artifact},
                     ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
