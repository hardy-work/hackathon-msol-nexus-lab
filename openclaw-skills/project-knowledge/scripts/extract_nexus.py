#!/usr/bin/env python3
"""Stage 2 extractor for the Nexus Plan workbook.

The workbook is a planning template with several independent tables.  This
extractor keeps every populated cell as a row-level record and separately
builds the small, deterministic person rollup used by the demo queries.
"""
import json
import re
from collections import Counter
from pathlib import Path

import openpyxl

from extract import cell_str, numeric_guard_ingest, scan_headers, scan_rows
from document_registry import current
from artifact_paths import artifact_path

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "raw"
DOC_ID = "nexus-plan"
DOC = current(DOC_ID)
VERSION = int(DOC["version"])
ORIGINAL = ROOT / str(DOC["original"])

ASSIGNEE = {
    "TùngDV": "tung-dv",
    "ĐôNT": "do-nt",
    "SơnBH": "son-bh",
    "KiênĐT": "kien-dt",
    "VinhNV": "vinh-nv",
    "LongVN": "long-vn",
    "HoàngMV": "hoang-mv",
}


def write_raw(raw_id, sheet, kind, body, payload):
    # Normalize generated Markdown so a rebuild is byte-stable.  Empty sheets
    # otherwise accumulate an extra blank line and appear as source changes.
    body = body.rstrip()
    md_path = artifact_path(ROOT, DOC, raw_id, "md")
    facts_path = artifact_path(ROOT, DOC, raw_id, "facts")
    md_path.parent.mkdir(parents=True, exist_ok=True)
    md_path.write_text(
        f"---\nraw_id: {raw_id}\ndoc_id: {DOC_ID}\nsheet: {sheet!r}\n"
        f"version: {VERSION}\nkind: {kind}\ngenerated_by: scripts/extract_nexus.py\n---\n\n"
        f"# {raw_id}\n\nNguồn: `{sheet}` trong `{DOC['original']}`\n\n{body}\n",
        encoding="utf-8")
    facts_path.write_text(
        json.dumps({"doc_id": DOC_ID, "version": VERSION, **payload},
                   ensure_ascii=False, indent=2), encoding="utf-8")


def rows_source(wb, raw_id, sheet, header_rows, start, cols=None, skip=None):
    ws = wb[sheet]
    if cols:
        a, b = cols.split(":")
        lo = openpyxl.utils.column_index_from_string(a)
        hi = openpyxl.utils.column_index_from_string(b)
    else:
        lo, hi = 1, ws.max_column or 1
    headers = scan_headers(ws, header_rows, hi, lo)
    rows, numeric = scan_rows(ws, sheet, headers, start, hi,
                              skip_rows=set(skip or []), min_col=lo)
    lines = [f"_Bảng `{sheet}` — {len(rows)} dòng dữ liệu._", ""]
    for rec in rows:
        lines.append(f"[r{rec['row']}] " + " | ".join(
            f"{c['header']}={c['value']}" for c in rec["cells"].values()))
    write_raw(raw_id, sheet, "rows", "\n".join(lines), {
        "kind": "rows", "sheet": sheet, "headers": headers,
        "rows": rows, "numeric_cells": numeric,
    })


def config_source(wb):
    ws = wb["Config"]
    ranges = {
        "provider": "E2:E15", "level": "F2:F15", "tech_stack": "G2:G15",
        "assignee": "H2:H15", "task_status": "I2:I15",
        "priority": "J2:J15", "role": "K2:K15",
    }
    vocabulary, cells = {}, []
    for field, ref in ranges.items():
        a, b = ref.split(":")
        c0, r0 = re.match(r"([A-Z]+)(\d+)", a).groups()
        _, r1 = re.match(r"([A-Z]+)(\d+)", b).groups()
        values = []
        for r in range(int(r0), int(r1) + 1):
            value = cell_str(ws[f"{c0}{r}"].value)
            if value and value not in values:
                values.append(value)
                cells.append({"field": field, "value": value,
                              "src": f"Config!{c0}{r}"})
        vocabulary[field] = values
    body = "\n".join([f"## {k}\n\n" + "\n".join(f"- `{v}`" for v in vs)
                      for k, vs in vocabulary.items()])
    write_raw("nexus-config", "Config", "vocabulary", body,
              {"vocabulary": vocabulary, "vocabulary_cells": cells})
    return vocabulary


def sprint_source(wb, vocab):
    ws = wb["Sprint 1"]
    headers = scan_headers(ws, [1, 2], 18, 1)
    headers.update({"H": "PLAN Start Date", "I": "PLAN End Date",
                    "K": "Actual Start Date", "L": "Actual End Date"})
    rows, numeric = scan_rows(ws, "Sprint 1", headers, 6, 18, min_col=1)
    facts, declared = {}, set(vocab["assignee"])
    observed = {"role": set(), "assignee": set(), "task_status": set(),
                "priority": set()}
    status_counts = Counter()
    priority_counts = Counter()
    remaining_total = 0.0
    for rec in rows:
        c = rec["cells"]
        label = cell_str(c.get("E", {}).get("value"))
        if not label:
            continue
        if label not in ASSIGNEE:
            raise ValueError(f"assignee chưa được curate trong mapping: {label}")
        slug = ASSIGNEE[label]
        node = facts.setdefault(slug, {"label": label, "roles": [],
                                       "task_count": 0, "estimate_h": 0.0,
                                       "actual_h": 0.0, "rows": []})
        node["task_count"] += 1
        node["rows"].append(rec["row"])
        role = cell_str(c.get("D", {}).get("value"))
        if role and role not in node["roles"]:
            node["roles"].append(role)
        for key, col in (("estimate_h", "G"), ("actual_h", "M")):
            raw = c.get(col, {}).get("value_num")
            if isinstance(raw, (int, float)):
                node[key] += float(raw)
        for dim, col in (("role", "D"), ("assignee", "E"),
                         ("priority", "F"), ("task_status", "Q")):
            value = cell_str(c.get(col, {}).get("value"))
            if value:
                observed[dim].add(value)
                if dim == "task_status":
                    status_counts[value] += 1
                elif dim == "priority":
                    priority_counts[value] += 1
        remaining = c.get("P", {}).get("value_num")
        if isinstance(remaining, (int, float)):
            remaining_total += float(remaining)
    for slug, node in facts.items():
        rr = node.pop("rows")
        span = f"Sprint 1!E{rr[0]}:E{rr[-1]}"
        node["task_count"] = numeric_guard_ingest(
            f"{slug}.task_count", node["task_count"], "task", span)
        node["estimate_h"] = numeric_guard_ingest(
            f"{slug}.estimate_h", round(node["estimate_h"], 4), "hour", span)
        node["actual_h"] = numeric_guard_ingest(
            f"{slug}.actual_h", round(node["actual_h"], 4), "hour", span)
    total_tasks = len([r for r in rows if cell_str(r["cells"].get("E", {}).get("value"))])
    total_estimate = sum(v["estimate_h"]["value"] for v in facts.values())
    total_actual = sum(v["actual_h"]["value"] for v in facts.values())
    body = "\n".join(["| assignee | role | task | estimate_h | actual_h |",
                      "|---|---|---:|---:|---:|"] + [
        f"| {v['label']} | {', '.join(v['roles'])} | {v['task_count']['value']} | "
        f"{v['estimate_h']['value']} | {v['actual_h']['value']} |"
        for v in facts.values()])
    summary_facts = {
        "task_count": numeric_guard_ingest(
            "sprint1.total_task_count", total_tasks, "task", "Sprint 1!A6:R65"),
        "estimate_h": numeric_guard_ingest(
            "sprint1.total_estimate_h", round(total_estimate, 4), "hour", "Sprint 1!A6:R65"),
        "actual_h": numeric_guard_ingest(
            "sprint1.total_actual_h", round(total_actual, 4), "hour", "Sprint 1!A6:R65"),
        "remaining_h": numeric_guard_ingest(
            "sprint1.total_remaining_h", round(remaining_total, 4), "hour", "Sprint 1!P6:P65"),
        "status_counts": {
            key: numeric_guard_ingest(
                f"sprint1.status.{key}", value, "task", "Sprint 1!Q6:Q65")
            for key, value in sorted(status_counts.items())
        },
        "priority_counts": {
            key: numeric_guard_ingest(
                f"sprint1.priority.{key}", value, "task", "Sprint 1!F6:F65")
            for key, value in sorted(priority_counts.items())
        },
    }
    payload = {"kind": "table", "sheet": "Sprint 1", "headers": headers,
               "rows": rows, "numeric_cells": numeric,
               "facts": facts, "observed_dimensions":
               {k: sorted(v) for k, v in observed.items()},
               "summary_facts": summary_facts}
    write_raw("nexus-sprint1", "Sprint 1", "table", body, payload)
    return facts


def rollup(vocab, sprint):
    facts = {}
    for label in vocab["assignee"]:
        slug = ASSIGNEE[label]
        if slug in sprint:
            src = sprint[slug]
            facts[slug] = {"label": label, "roles": src["roles"],
                           "in_config": True, "appears_in_sprints": True,
                           "task_count": src["task_count"],
                           "estimate_h": src["estimate_h"],
                           "actual_h": src["actual_h"]}
        else:
            facts[slug] = {"label": label, "roles": [], "in_config": True,
                           "appears_in_sprints": False,
                           "task_count": numeric_guard_ingest(
                               f"{slug}.task_count", 0, "task", "Config!H2:H15"),
                           "estimate_h": numeric_guard_ingest(
                               f"{slug}.estimate_h", 0, "hour", "Config!H2:H15"),
                           "actual_h": numeric_guard_ingest(
                               f"{slug}.actual_h", 0, "hour", "Config!H2:H15")}
    body = "\n".join(["| người | vai trò | task | estimate_h | actual_h |",
                      "|---|---|---:|---:|---:|"] + [
        f"| {v['label']} | {', '.join(v['roles']) or '—'} | "
        f"{v['task_count']['value']} | {v['estimate_h']['value']} | {v['actual_h']['value']} |"
        for v in facts.values()])
    write_raw("nexus-people", "(rollup Sprint 1)", "rollup", body,
              {"facts": facts, "zero_task": [v["label"] for v in facts.values()
                                               if not v["appears_in_sprints"]]})


def main():
    wb = openpyxl.load_workbook(ORIGINAL, data_only=True)
    RAW.mkdir(exist_ok=True)
    vocab = config_source(wb)
    rows_source(wb, "nexus-resource-plan", "Resource plan", [5, 6, 7], 8, "A:AK")
    rows_source(wb, "nexus-summary", "Summary project", [3], 4, "B:K")
    rows_source(wb, "nexus-master-schedule", "Master schedule", [2, 3], 5, "A:N")
    rows_source(wb, "nexus-backlog", "Backlog", [1], 2, "A:H")
    rows_source(wb, "nexus-risk", "Risk management ", [1], 3, "A:H", skip=[2])
    rows_source(wb, "nexus-issue", "Isssue management", [1], 3, "A:H", skip=[2])
    sprint = sprint_source(wb, vocab)
    rollup(vocab, sprint)
    print("✓ Nexus extraction: 8 sheet routes, mẫu R-000/I-000 đã loại")


if __name__ == "__main__":
    main()
