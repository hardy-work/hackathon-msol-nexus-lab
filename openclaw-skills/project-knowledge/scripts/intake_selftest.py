#!/usr/bin/env python3
"""Offline tests for automatic file identity and version registration."""
from __future__ import annotations

import hashlib
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl
import yaml
from openpyxl.styles import PatternFill

import document_registry
import intake


def digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def workbook(path: Path, value: str, styled: bool = False) -> None:
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Plan"
    sheet["A1"] = "Task"
    sheet["B1"] = value
    if styled:
        sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    book.save(path)


def write_registry(root: Path, original: Path, doc_id: str = "plan", version: int = 1) -> None:
    payload = {"documents": [{
        "doc_id": doc_id,
        "version": version,
        "original": original.relative_to(root).as_posix(),
        "kind": "xlsx",
        "sha256": digest(original),
        "status": "canonical",
        "current": True,
        "supersedes": None,
        "visibility": "internal",
        "extractor": "nexus",
        "raw_paths": ["raw/plan.md"],
    }]}
    (root / "documents.yml").write_text(
        yaml.safe_dump(payload, allow_unicode=True, sort_keys=False), encoding="utf-8")


def main() -> int:
    with tempfile.TemporaryDirectory(prefix="pk-intake-") as temp:
        root = Path(temp)
        (root / "originals").mkdir()
        (root / "raw").mkdir()
        current = root / "originals/plan.xlsx"
        formatted = root / "uploads/Plan.xlsx"
        updated = root / "uploads/Plan v2.xlsx"
        current.parent.mkdir(exist_ok=True)
        formatted.parent.mkdir(exist_ok=True)
        workbook(current, "old")
        workbook(formatted, "old", styled=True)
        workbook(updated, "new")
        write_registry(root, current)

        review = intake.decide(root, formatted)
        assert review["flow"] == "identity_review"
        assert review["candidate_doc_ids"] == ["plan"]

        no_op = intake.decide(root, formatted, confirmed_doc_id="plan")
        assert no_op["flow"] == "no_op"

        review = intake.decide(root, updated)
        assert review["flow"] == "identity_review"
        decision = intake.decide(root, updated, confirmed_doc_id="plan")
        assert decision["flow"] == "reingest"
        assert decision["doc_id"] == "plan"
        assert decision["from_version"] == 1
        assert decision["to_version"] == 2
        registered = intake.register(root, updated, decision)
        assert registered["registered_original"] == "originals/plan@v2.xlsx"
        assert (root / registered["registered_original"]).exists()
        assert document_registry.current("plan", root)["version"] == 2
        assert document_registry.by_version("plan", 1, root)["current"] is False
        assert document_registry.by_version("plan", 2, root)["raw_paths"] == [
            "raw/plan@v2.md"
        ]
        assert intake.decide(root, updated)["flow"] == "duplicate"

        try:
            intake.register(intake.ROOT, updated, decision)
        except ValueError as exc:
            assert "canonical skill root" in str(exc)
            assert "staging/worktree" in str(exc)
        else:
            raise AssertionError("register() không được ghi vào canonical root")

        cli = subprocess.run(
            [sys.executable, str(Path(intake.__file__)), "--file", str(updated),
             "--doc-id", "plan", "--apply"],
            capture_output=True, text=True, check=False,
        )
        assert cli.returncode == 2
        assert "--root" in cli.stderr
        assert "staging/worktree" in cli.stderr

    with tempfile.TemporaryDirectory(prefix="pk-intake-new-") as temp:
        root = Path(temp)
        (root / "originals").mkdir()
        source = root / "uploads/New Plan.xlsx"
        source.parent.mkdir()
        workbook(source, "first")
        (root / "documents.yml").write_text("documents: []\n", encoding="utf-8")
        decision = intake.decide(root, source)
        assert decision["flow"] == "initial_ingest"
        assert decision["doc_id"].startswith("new-plan-")
        assert decision["source_name"] == "New Plan.xlsx"
        assert decision["kind"] == "xlsx"
        assert decision["version"] == 1
        intake.register(root, source, decision)
        registered_new = document_registry.current(decision["doc_id"], root)
        assert registered_new["doc_id"] == decision["doc_id"]
        assert registered_new["version"] == 1
        assert registered_new["extractor"] == "spreadsheet"
        assert registered_new["raw_paths"] == [f"raw/{decision['doc_id']}.md"]

    print("✓ intake self-test: 10/10 qua")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
